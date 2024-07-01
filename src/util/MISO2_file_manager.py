#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Collection of objects and functions for dealing with ODYM/MISO2 input sets.

@author: bgrammer
"""

import re
import logging as logger
import glob
import os
import openpyxl
import pandas as pd
import numpy as np
from preprocessing.util import MISO2_input_constants as InputConstants
from preprocessing.util.MISO2_file_types import MISO2FileTypes


class MISO2FileManager:
    """
    Manages paths and filenames associated with a single ODYM/MISO2 input set.

    Additional data, uncertainty values and input sources are optional.
    """
    def __init__(self, data_dir, config_filename, additional_data_dir=None,
                 uncertainty_values_dir=None, input_sources_dir=None):
        parameter_filenames = get_parameter_filenames(data_dir, config_filename, "all")

        self.file_paths = {MISO2FileTypes.DATA_DIR: data_dir,
                           MISO2FileTypes.CONFIG: os.path.join(data_dir, config_filename),
                           MISO2FileTypes.UNCERTAINTY_DISTRIBUTION: os.path.join(
                               data_dir, InputConstants.UNCERTAINTY_DISTRIBUTION_FILENAME),
                           MISO2FileTypes.CLASSIFICATION: os.path.join(
                               data_dir, InputConstants.CLASSIFICATION_FILENAME),
                           MISO2FileTypes.PARAMETER_FILE_LIST: parameter_filenames["list"],
                           MISO2FileTypes.PARAMETER_FILE_DICT: parameter_filenames["dict"]
                           }

        if additional_data_dir:
            self.file_paths[MISO2FileTypes.ADDITIONAL_DATA_DIR] = additional_data_dir
            self.file_paths[MISO2FileTypes.COUNTRY_REGION_CORRESPONDENCE] = os.path.join(
                additional_data_dir, InputConstants.COUNTRY_CORRESPONDENCE_FILENAME)
            self.file_paths[MISO2FileTypes.CV_LOOKUP] = os.path.join(
                additional_data_dir, InputConstants.UNCERTAINTY_SCORE_LOOKUP_FILENAME)
            self.file_paths[MISO2FileTypes.METADATA_TEMPLATE] = os.path.join(
                additional_data_dir, InputConstants.METADATA_FILENAME)
        else:
            self.file_paths[MISO2FileTypes.ADDITIONAL_DATA_DIR] = None
            self.file_paths[MISO2FileTypes.COUNTRY_REGION_CORRESPONDENCE] = None
            self.file_paths[MISO2FileTypes.CV_LOOKUP] = None

        if uncertainty_values_dir:
            self.file_paths[MISO2FileTypes.UNCERTAINTY_VALUES_DIR] = uncertainty_values_dir
        else:
            self.file_paths[MISO2FileTypes.UNCERTAINTY_VALUES_DIR] = None

        if input_sources_dir:
            self.file_paths[MISO2FileTypes.INPUT_SOURCES_DIR] = input_sources_dir
        else:
            self.file_paths[MISO2FileTypes.INPUT_SOURCES_DIR] = None

    def load_parameters(self, parameters=None, sheet_name=InputConstants.VALUES_SHEET_NAME):
        """
        Returns a dictionary of parameter values.
        Args:
            parameters(list): List of parameters to read. If None (default), all will be read.
        Returns:
             parameter_dict(dict): Dict with parameter name as key and values sheet as pd.DataFrame.
        """

        parameter_dict = {}
        for parameter_name, file_name in self.file_paths[MISO2FileTypes.PARAMETER_FILE_DICT].items():
            if parameters is None:
                read_in = True
            elif parameter_name in parameters:
                read_in = True
            else:
                read_in = False

            if read_in:
                try:
                    logger.info(f"Reading {parameter_name} from {file_name}")
                    df = pd.read_excel(
                        os.path.join(self.get_filepath(MISO2FileTypes.DATA_DIR), file_name),
                        sheet_name=sheet_name)
                    if InputConstants.G_INDEX in df.columns:
                        df = df.set_index(InputConstants.REMGT_INDEX)
                    else:
                        df = df.set_index(InputConstants.REMT_INDEX)
                    df.columns = [int(entry) for entry in list(df.columns)]

                    parameter_dict[parameter_name] = df
                except Exception as e:
                    logger.error(f"Error trying to load parameter file: {repr(e)}")

        return parameter_dict

    def load_sources(self, sources=None):
        """
        Parses all xlsx files from a directory or list of paths to create a dictionary of uncertainty inputs.

        Args:
            sources(str/list/None): Path to uncertainty sources or list of paths to uncertainty source files. \
            If None (default), the value will be taken from the input parser.

        Returns:
            parameter_uncertainty_dataframes(dict): Dictionary with parameter names as keys \
                and a list of pd.DataFrames as values.
        """
        logger.info(f"Parsing source tables {sources}")
        if sources is None:
            sources = self.get_filepath(MISO2FileTypes.INPUT_SOURCES_DIR)

        if isinstance(sources, str):
            uncertainties_glob = os.path.join(sources, "*.xlsx")
            uncertainties_files = glob.glob(uncertainties_glob)
            for file in uncertainties_files:
                # removing temp files, don't judge
                if "~$" in file:
                    uncertainties_files.remove(file)

            logger.info(f"Found uncertainty files: {uncertainties_files}")
        elif isinstance(sources, list):
            uncertainties_files = sources
        else:
            error_msg = f"Received unknown argument to parsing function: {type(sources)}"
            logger.error(error_msg)
            raise ValueError(error_msg)

        return _parse_source_tables(uncertainties_files)

    def load_xls(self, path_type, sheet=None):
        """
        Load XLS of desired path_type.
        """
        _check_for_filetype(path_type)
        try:
            file_path = self.get_filepath(path_type)
            file = pd.read_excel(file_path, sheet_name=sheet)
            return file

        except (IOError, ValueError) as error:
            error_msg = f"Could not load {path_type} due to an error."
            logger.error(error_msg)
            logger.exception(error)
            raise

    def get_filepath(self, path_type):
        """
        Returns filepaths for appropriate MISO2FileType enum.

        Args:
            path_type(MISO2FileType): Enum of available file types.
        Returns:
            file_path(str/list/dict): Filepath as string or collection of filepaths as string.
        Raises:
            KeyError: When FileType is not available.
        """
        _check_for_filetype(path_type)
        try:
            file_path = self.file_paths[path_type]
            return file_path
        except KeyError as exc:
            error_msg = f"Could not find path of type {path_type} in path dictionary. \
                        Either the argument is misspelled or the input parser was not set up to contain this data."
            logger.error(error_msg)
            raise KeyError(error_msg) from exc

    def update_filepath(self, path_type, new_value):
        """
        Set new value for filepath.

        Args:
            path_type(MISO2FileType): FileType to be specified.
            new_value(str): New path.
        """
        _check_for_filetype(path_type)
        try:
            if path_type in self.file_paths:
                self.file_paths[path_type] = new_value
                logger.info(f"Updated {path_type} with {new_value}")
        except KeyError as exc:
            error_msg = f"Updating path of type {path_type} failed, as key not present in path dictionary."
            logger.error(error_msg)
            raise KeyError(error_msg) from exc

    def write_uncertainties(self, material, new_uncertainties_dict, mode):
        """
        Write uncertainties back to uncertainty files.

        Filepaths are gathered from object class attributes. In mode overwrite, original values \
            (if any) are replaced. In update mode, only values present in the original value will \
                be updated with new values generated from sources. In "update_append", present values \
                    will be updated and those only present in sources are appended.

        In update append mode, the number of columns between existing and new values must match.

        Args:
            material(str): Name of material.
            new_uncertainties_dict(dict): Dict of pd.DataFrames with parameter name as key.
            mode(str): Update mode, one of "overwrite", "update" or "update_append.
        """
        logger.info(f"Writing uncertainties back to file in mode {mode}")
        uncertainty_values_dir = self.get_filepath(MISO2FileTypes.UNCERTAINTY_VALUES_DIR)

        uncertainties_files = glob.glob(os.path.join(uncertainty_values_dir, "*.xlsx"))
        for file in uncertainties_files:
            # removing temp files, don't judge
            if "~$" in file:
                uncertainties_files.remove(file)

        uncertainties_files_by_material = {}

        for file_name in uncertainties_files:
            material_name = regex_material_name(file_name)
            if material_name is not None:
                uncertainties_files_by_material[material_name] = file_name
            # this can be replaced with access to material column

        uncerts_filename = uncertainties_files_by_material[material]
        original_uncerts_sheets = pd.read_excel(uncerts_filename, sheet_name=None)
        parameter = "Could not read any parameters"
        try:
            for parameter, new_uncerts in new_uncertainties_dict.items():
                logger.info(f"Writing {parameter}")

                index_check(new_uncerts)

                if parameter in original_uncerts_sheets:
                    original_uncerts = original_uncerts_sheets[parameter]
                    original_uncerts = original_uncerts.set_index(new_uncerts.index.names, drop=True).sort_index()
                    updated_uncerts = _update_existing_values(original_uncerts, new_uncerts, mode)
                else:
                    updated_uncerts = new_uncerts

                # catch case that parameter longer than 32 chars (!)
                if len(parameter) > 31:
                    sheet_name = parameter[:31]
                else:
                    sheet_name = parameter

                with pd.ExcelWriter(uncerts_filename, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    updated_uncerts.reset_index().to_excel(writer, sheet_name=sheet_name, index=False)

        except (IOError, ValueError) as error:
            error_msg = f"Error trying to write output for {material} and {parameter} \n"
            logger.error(error_msg)
            logger.exception(error)

    def write_values(self, parameter, parameter_dfs, mode, sheet_name=InputConstants.VALUES_SHEET_NAME):
        """
        Write values back to parameter files.

        Filepaths are gathered from object class attributes.

        Args:
            parameter(str): Name of parameter
            parameter_dfs(list): List of pd.DataFrames
            mode(str): Update mode, one of "overwrite", "update" or "update_append".
        """
        logger.info(f"Writing values for {parameter} with mode: {mode}")
        data_dir = self.get_filepath(MISO2FileTypes.DATA_DIR)
        parameter_filenames_dict = self.get_filepath(MISO2FileTypes.PARAMETER_FILE_DICT)

        index_check(parameter_dfs)

        try:
            values_file = os.path.join(data_dir, parameter_filenames_dict[parameter])
            original_df = pd.read_excel(values_file, sheet_name=None)
            new_values = pd.concat(parameter_dfs).sort_index()
            if sheet_name in original_df:
                original_values = original_df[sheet_name]
                original_values = original_values.set_index(new_values.index.names, drop=True)
                original_values = original_values.sort_index()
                updated_values = _update_existing_values(original_values, new_values, mode)
            else:
                logger.warning(f"Could not find a {sheet_name} sheet in loaded excel workbook."
                               f"Writing back new values to new sheet")
                updated_values = new_values
            with pd.ExcelWriter(values_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                updated_values.reset_index().to_excel(writer, sheet_name=sheet_name, index=False)

        except (IOError, ValueError) as error:
            error_msg = f"Error trying to write output for {parameter} \n" + repr(error)
            logger.error(error_msg)
            logger.exception(error)
        except KeyError as key_error:
            error_msg = f"Error accessing key for {parameter}"
            logger.error(error_msg)
            logger.exception(key_error)
            raise key_error


def _update_existing_values(original_values, new_values, mode):
    """
    Update existing values for input parameters.

    Args:
        original_values(pd.DataFrame): Old values.
        new_values(pd.DataFrame): New values.
        mode(str): One of overwrite, update or update-append.
    """

    if original_values.empty or mode == "overwrite":
        return new_values
    if mode == "update":
        original_values.update(new_values)
        return original_values
    if mode == "update_append":
        if not original_values.columns.equals(new_values.columns):
            raise ValueError("While updating, columns of original and new values \
                                         were found not to match")
        index_intersection = original_values.index.intersection(new_values.index)
        if index_intersection.equals(original_values.index):
            original_values.update(new_values)
            updated_values = original_values
        elif index_intersection.empty:
            updated_values = pd.concat([original_values, new_values]).sort_index()
        else:
            difference_new_values = new_values.index.difference(original_values.index)
            original_values.update(new_values)
            updated_values = pd.concat([original_values, new_values.loc[difference_new_values]]).sort_index()
        return updated_values

    error_msg = f"Unknown argument combination in type {mode}. This can happen if you try to update \
                    non-existing data"
    logger.error(error_msg)
    raise ValueError(error_msg)


def _parse_source_tables(source_files):
    """
    Parse source files into dictionary.

    Args:
        source_files(list): List of filepaths.
    Returns:
        parameter_source_dataframes(dict): Dictionary of nested dicts with parameters as keys and materials as keys for
        pd.DataFrames of values.
    """
    parameter_source_dataframes = {}

    for file in source_files:
        try:
            sources = pd.read_excel(file, sheet_name=None)
            for sheet_name, dataframe in sources.items():
                material_by_file = np.unique(dataframe["Material"])
                parameter_by_file = np.unique(dataframe["Parameter"])
                # dataframes should only have one material and parameter per sheet

                if len(material_by_file) == 1 and len(parameter_by_file) == 1:
                    material_by_file = material_by_file[0]
                    parameter_by_file = parameter_by_file[0]
                else:
                    error_msg = f"Received dataframe sheet {sheet_name} with more than one parameter name \
                            in columns: {material_by_file} from file {file}"
                    logger.error(error_msg)
                    raise ValueError(error_msg)
                if material_by_file in parameter_source_dataframes:
                    parameter_source_dataframes[material_by_file][parameter_by_file] = dataframe
                else:
                    parameter_source_dataframes[material_by_file] = {parameter_by_file: dataframe}
        except (ValueError, TypeError) as error:
            logger.error(f"Error for file: {file} with sheet name {sheet_name} and df {dataframe.head()}")
            logger.exception(error)
    if len(parameter_source_dataframes) == 0:
        error_msg = f"Warning, no uncertainty files were parsed. Did you point to the right directory: " \
                    f" {source_files}?"
        logger.error(error_msg)
        raise ValueError(error_msg)
    return parameter_source_dataframes


def get_parameter_filenames(data_path, config_file, return_type="list"):
    """
    Helper function that generates correct parameter filenames from config.

    This function uses reworked code from ODYM's parser function.

    Args:
        data_path(str): Path to config file.
        config_file(str): Name of ODYM config file.
        return_type(str): Format to return filenames in. One of "list" (default), "dict" or "all". "All" returns a dict
        with both list and key as values.

    Returns:
        file_names(list/dict): Parameter input file names. If dict, the parameter name without version number \
            is provided as key.
    """

    model_config_sheet = openpyxl.load_workbook(os.path.join(data_path, config_file)).active

    p_lix = 0
    file_names = {}

    while True:
        if model_config_sheet.cell(p_lix+1, 2).value == 'Model Parameters':
            break
        p_lix += 1

    p_lix += 2
    while model_config_sheet.cell(p_lix+1, 3).value is not None:
        name = model_config_sheet.cell(p_lix+1, 3).value
        version = model_config_sheet.cell(p_lix+1, 5).value
        file_names[name] = name + "_" + version + ".xlsx"
        p_lix += 1

    if return_type == "list":
        return file_names.values()
    if return_type == "dict":
        return file_names
    if return_type == "all":
        all_types = {"list": file_names.values(), "dict": file_names}
        return all_types

    error_msg = f"Illegal argument: {return_type}"
    logger.error(error_msg)
    raise ValueError(error_msg)


def regex_material_name(input_string):
    """
    Split file path to uncertainty parameter file with .xls ending into the material name.
    Filenames must be of format uncertainty_<material_name>.xlsx

    Args:
        input_string(str): Path to regex for material name.

    Returns:
        name(str): Material name.
    """
    pattern = r'(?<=uncertainty_).+(?=.xls)'
    split_string = input_string.split("/")[-1]
    name = re.search(pattern, split_string)

    if name is None:
        logger.warning(f"Could not create a material name out of filename: {input_string}")
        return None

    return name.group()


def _check_for_filetype(test_object):
    """
    Filetype check. Could be moved to a decorator.

    Args:
        test_object(MISO2FileTypes): Should be a filetype.
    Raises:
        ValueError: When incorrect object is supplied.
    """
    if not isinstance(test_object, MISO2FileTypes):
        error_msg = f"Invalid argument path type: {type(test_object)}"
        logger.error(error_msg)
        raise ValueError(error_msg)


def index_check(dfs):
    """
    Check dataframes or a list of dataframes if they have a valid REMT / REMGT index.

    dfs(pd.DataFrame/list): List of or single pd.DataFrame

    Raises:
        ValueError: If index is different than expected.
    """
    valid_indices = [InputConstants.REMT_INDEX, InputConstants.REMGT_INDEX]

    if isinstance(dfs, list):
        for df in dfs:
            if df.index.names not in valid_indices:
                raise ValueError(f"Not a valid REMT / REMGT index: {df.index.names} ")
    elif isinstance(dfs, pd.DataFrame):
        if dfs.index.names not in valid_indices:
            raise ValueError(f"Not a valid REMT / REMGT index: {dfs.index.names} ")
    else:
        raise ValueError(f"Illegal argument type to function: {type(dfs)}")
