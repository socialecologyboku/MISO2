#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Collection and wrappers for input data sanity check.

Created on Mon Nov  7 14:29:28 2022

@author: bgrammer

"""

import logging as logger
import os
import pandas as pd
from openpyxl.reader.excel import load_workbook

from preprocessing.util import MISO2_input_constants as InputConstants


def input_data_sanity_check(data_dir, file_names, write_back=True):
    """
    This script should be executed after updating the global config and uncertainties.

    It checks for the occurrence of materials, countries, elements and enduse sector in respect to all input parameter
    files and the classification properties file and will print out any differences (e.g. if any material is present
    in the classification, but not in the data, and vice versa). The script will check for the occurrence of negative
    and nan values. Finally, it will check that values and uncertainty values match, and then sort all by index.
    These changes can be written back to the input files if desired.

    An additional check is in place to ensure the number of data rows and columns in values is congruent with the
    cover sheet. This will NOT be fixed automatically.

    Args:
        data_dir(str): Path to directory with src input data.
        file_names(list): Parameter input file names.
        write_back(bool): Write back data frames in a sorted order. Defaults to true.
    """

    classifications = pd.read_excel(os.path.join(
        data_dir, InputConstants.CLASSIFICATION_FILENAME), sheet_name=None)

    parameters = read_parameters(data_dir, file_names)
    check_sheet_size(parameters)
    sort_dfs_and_write_back(parameters, write_back=write_back)
    check_values_against_classification(parameters, classifications)
    check_negative_values(parameters)
    check_nan_values(parameters)


def sort_input_by_index(dataframe):
    """
    Sets either REMT or REMGT Index depending on column names.

    Args:
        dataframe(pd.DataFrame): Dataframe without index.

    Returns:
         dataframe(pd.DataFrame): Dataframe with index.
    """
    if InputConstants.G_INDEX in dataframe.columns:
        index = InputConstants.REMGT_INDEX
    else:
        index = InputConstants.REMT_INDEX

    dataframe.set_index(index, inplace=True, drop=False)
    dataframe = dataframe.sort_index()

    return dataframe


def read_parameters(data_dir, file_names):
    """
    Read in parameter files from path.

    Args:
        data_dir(str): Data directory
        file_names(list): List of parameter filenames (with correct extensions)

    Returns:
         parameters(dict): Dictionary with filepath as keys and dicts of pd.DataFrames as values.
    """
    file_paths = [os.path.join(data_dir, name) for name in file_names]
    parameters = {}
    for file_path in file_paths:
        parameters[file_path] = pd.read_excel(file_path, sheet_name=None)

    return parameters


def check_values_against_classification(parameters, classifications, direction="both"):
    """
    Check presence of values against all dimensions of classifications. Results are written to log.
    Checks are skipped for empty dataframes

    Args:
        parameters(dict): Dictionary of parameters and pd.DataFrames
        classifications(dict): Dictionary of classification dataframes
        direction(str): Check direction of presence, one of "classification", "dataframe" or "both" (default).
    """
    def check_values(parameter_name, dataframe, classification, column_name, check_direction):
        df_values_from_column = set(dataframe.reset_index()[column_name])

        class_set = set(classification)

        missing_df = class_set.difference(df_values_from_column)
        missing_class = df_values_from_column.difference(classification)

        if missing_df and check_direction in ["both", "dataframe"]:
            logger.warning(f"Input data parameter {parameter_name} misses values of type {column_name} which "
                           f"are present in classification file: {missing_df} \n")
        if missing_class and check_direction in ["both", "classification"]:
            logger.error(f"Classification file misses values from parameter {parameter_name} of "
                         f"type {column_name} that are present in input values: {missing_class} \n")

    materials = classifications[InputConstants.ODYM_MATERIAL[0]]["Item"]
    end_use_sectors = classifications[InputConstants.ODYM_END_USE[0]]["Item"]
    countries = classifications[InputConstants.ODYM_REGION[0]]["Item"]
    elements = classifications[InputConstants.ODYM_ELEMENT[0]]["Item"]

    for parameter, df in parameters.items():
        if isinstance(df, dict):
            df_values = df[InputConstants.VALUES_SHEET_NAME]
        elif isinstance(df, pd.DataFrame):
            df_values = df
        else:
            raise ValueError(f"Type: {type(df)} not processable. Must be either dictionary or pd.DataFrame")
        if df_values.size == 0:
            logger.warning(f"Dataframe is empty for {parameter}, skipping further checks")
            continue

        check_values(parameter, df_values, countries, InputConstants.R_INDEX, direction)

        check_values(parameter, df_values, materials, InputConstants.M_INDEX, direction)

        check_values(parameter, df_values, elements, InputConstants.E_INDEX, direction)

        if InputConstants.G_INDEX in df_values.columns:
            check_values(parameter, df_values, end_use_sectors, InputConstants.G_INDEX, direction)


def sort_dfs_and_write_back(parameters, write_back=False):
    """
    Modifies dataframes in given dict in place by sorting.

    Args:
        parameters(dict): Dict of pd.DataFrames.
        write_back(bool): Write back values to files immediately. Defaults to False

    """
    for file_name, df in parameters.items():
        try:
            if InputConstants.UNCERTAINTY_SHEET_NAME in df.keys():
                sorted_values = sort_input_by_index(df[InputConstants.VALUES_SHEET_NAME].copy())
                sorted_uncerts = sort_input_by_index(df[InputConstants.UNCERTAINTY_SHEET_NAME].copy())

                sorted_values.reset_index(inplace=True, drop=True)
                sorted_uncerts.reset_index(inplace=True, drop=True)
                # reset index before writing back
                assert sorted_values.shape == sorted_uncerts.shape, \
                    f"Shapes of df not equal for parameter: {file_name}"

                assert sorted_values.index.equals(sorted_uncerts.index), \
                    f"Index not equal for parameter: {file_name}"

            else:
                sorted_uncerts = None
                sorted_values = sort_input_by_index(df[InputConstants.VALUES_SHEET_NAME].copy())
                sorted_values.reset_index(inplace=True, drop=True)

            if write_back:
                with pd.ExcelWriter(file_name, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    sorted_values.to_excel(writer, sheet_name=InputConstants.VALUES_SHEET_NAME, index=False)
                    if isinstance(sorted_uncerts, pd.DataFrame):
                        sorted_uncerts.to_excel(writer, sheet_name=InputConstants.UNCERTAINTY_SHEET_NAME, index=False)
        except (KeyError, IOError) as error:
            logger.error("Error during sorting and writing dataframes")
            logger.exception(error)
        except AssertionError as error:
            logger.error("Unsorted values found")
            logger.exception(error)


def check_negative_values(parameters):
    """
    Checks for number of negative values in a dataframe.

    Args:
        parameters(dict): Dictionary of pd.DataFrames corresponding to sheets of ODYM input file.

    """
    for file_name, df_dict in parameters.items():
        df = df_dict[InputConstants.VALUES_SHEET_NAME]
        df = df.apply(pd.to_numeric, errors="coerce")
        df = df[df.notnull()].dropna(axis=1)
        negs = df.values[df.values < 0]

        if negs.size > 0:
            logger.warning(f"Parameter {file_name} contains {negs.size} negative values \n")


def check_nan_values(parameters):
    """
    Checks for Nan values in dataframe.

    Args:
        parameters(dict): Dictionary of pd.DataFrames corresponding to sheets of ODYM input file.

    """
    nan_values = {}
    for file_name, df_dict in parameters.items():
        df = df_dict[InputConstants.VALUES_SHEET_NAME]
        nans = df.isnull().sum().sum()
        if nans > 0:
            logger.warning(f"Parameter {file_name} contains {nans} nan values \n")
            nan_values[file_name] = df[df.isnull().any(axis=1)]

    return nan_values


def check_sheet_size(parameters, write_back=False):
    """
    Check if sheet size given on ODYM input file corresponds to actual number of rows and columns.

    The function logs to warnings.

    Args:
        parameters(dict): Dictionary of pd.DataFrames corresponding to sheets of ODYM input file.
        write_back(bool): Write back corrected values, if any.

    """
    for file_name, dfs in parameters.items():
        try:
            cover_sheet = dfs[InputConstants.COVER_SHEET_NAME]
            index = "MISO2 Parameter File"
            # this should be the "MISO2 Parameter file" or similar header of the cover sheet
            new_cover_sheet = cover_sheet.set_index(index)
            drt = new_cover_sheet.loc["Dataset_RecordType"]

            last_item = None
            no_rows = 0
            no_cols = 0
            for series_tuple in drt.items():
                if last_item == "No_Rows":
                    no_rows = series_tuple[1]
                elif last_item == "No_Cols":
                    no_cols = series_tuple[1]
                last_item = series_tuple[1]
            values = dfs[InputConstants.VALUES_SHEET_NAME]
            data_rows = values.shape[0]
            index_cols = set(InputConstants.REMGT_INDEX)
            data_cols = len(set(values.columns).difference(index_cols))

            if no_rows != data_rows:
                logger.warning(f"For {file_name} Data values rows do not match cover sheet: {no_rows} vs {data_rows}")
            if no_rows < data_rows:
                logger.warning("WARNING: Number of rows in cover sheet is smaller than number of data rows, \
                      some values will not be parsed")
            if no_cols != data_cols:
                logger.warning(f"For {file_name} Data values cols do not match cover sheet: {no_cols} vs {data_cols}")
            if no_cols < data_cols:
                logger.warning("WARNING: Number of cols in cover sheet is smaller than number of data cols, \
                      some values will not be parsed")

            if (no_rows != data_rows or no_cols != data_cols) and write_back:
                update_cover_sheet(file_name, InputConstants.COVER_SHEET_NAME, data_rows, data_cols)

        except (AttributeError, KeyError) as error:
            logger.error(f"Error during check of parameter {file_name}")
            logger.exception(error)


def update_cover_sheet(file_path, sheet_name, no_rows, no_cols):
    logger.info(f"Updating {file_path} and saving back to file")
    workbook = load_workbook(filename=file_path)
    worksheet = workbook[sheet_name]

    for row in worksheet.iter_rows(min_row=1, max_col=worksheet.max_column, max_row=worksheet.max_row, values_only=False):
        for cell in row:
            if cell.value == "No_Rows":
                # Assuming the value to update is in the next column
                next_cell = worksheet.cell(row=cell.row, column=cell.column + 1)
                next_cell.value = no_rows
            elif cell.value == "No_Cols":
                next_cell = worksheet.cell(row=cell.row, column=cell.column + 1)
                next_cell.value = no_cols

    workbook.save(filename=file_path)
    workbook.close()

