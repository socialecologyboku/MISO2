#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Feb 15 14:03:01 2022

@author: jstreeck & bgrammer
"""
import glob
from importlib.metadata import version
import shutil
from copy import deepcopy
import logging as logger
import uuid
import pickle
import os
import openpyxl
import numpy as np
import pandas as pd
import odym.ODYM_Functions as msf
import odym.ODYM_Classes as msc
import preprocessing.checks.MISO2_checks
import preprocessing.util.MISO2_input_constants as InputConstants
from model.config.MISO2_sensitivity_by_parameter import MISO2SensitivityByParameter
from util.MISO2_exceptions import WasteRateError, InputLimitError
from model.config.MISO2_monte_carlo import MISO2MonteCarlo, transform_stats_array_to_dicts
from model.config.MISO2_monte_carlo_systematic import MISO2MonteCarloSystematic
from model.config.MISO2_monte_carlo_systematic_v2 import MISO2MonteCarloSystematic_v2
import util.MISO2_functions as misof
from datetime import datetime


class MISO2Config:
    """
    Class that holds read-only parameters and data for the MISO2 model.

    Input data is parsed from .xls files and brought into a format accepted by the ODYM module.
    Created with a unique identifier to allow for pickling and reloading the config.
    The config object is necessary to construct survival functions. If injected into a MISO2 \
        model object, it will a pass a copy of the input data for computation. When it parses data with uncertainty,
    Monte Carlo randomization of the data can be enabled.

    Args:
        config_path(str): Default save path when config is pickled.
        data_path(str): Path to input data.
        timestamp(str): Time of config creation. Will be used for default filenames of logs and output.
        mc_batch_size(int): Number of randomized copies of the data that are created when Monte Carlo is active.

    Attributes:
        config_path(str): A str that gives the default save location for config.

    """

    __slots__ = "config_path", "data_path", "timestamp", "model_classification", \
        "parameter_dict", "unique_id", "_monte_carlo", "script_config", "endUse_aggregates", \
        "master_classification", "index_table", "index_table_classification_names", \
        "model_duration", "model_time_start", "model_time_end", "IT_Aspects", \
        "IT_Description", "IT_Dimension", "IT_Classification", "IT_Selector", \
        "IT_IndexLetter", "PL_Names", "PL_Description", "PL_Version", \
        "PL_IndexStructure", "PL_IndexMatch", "PL_IndexLayer", \
        "PrL_Number", "PrL_Name", "PrL_Comment", "PrL_Type", "Nt", \
        "Ne", "Nc", "Nm", "Nr", "Ng", "selectors", "material_position", \
        "multiplier_cementBitumen", "uncertainty_settings", "nr_start", "nr_stop", \
        "mc_batch_size", "new_scrap_cycle", "sector_position", "sensitivity_settings", "new_scrap_cycle_cutoff"

    def __init__(self, config_path, data_path, timestamp, mc_batch_size=500):
        """U
        Create a MISO Config object.
        """
        self.mc_batch_size = mc_batch_size
        self.config_path = config_path
        self.data_path = data_path
        self.timestamp = timestamp
        self.model_classification = {}
        self.parameter_dict = {}
        self.uncertainty_settings = None
        self.unique_id = uuid.uuid4()
        self._monte_carlo = None

        self.script_config = None
        self.master_classification = None
        self.index_table = None
        self.index_table_classification_names = None
        self.model_duration = None
        self.model_time_start = None
        self.model_time_end = None

        self.IT_Aspects = None
        self.IT_Description = None
        self.IT_Dimension = None
        self.IT_Classification = None
        self.IT_Selector = None
        self.IT_IndexLetter = None
        self.PL_Names = None
        self.PL_Description = None
        self.PL_Version = None
        self.PL_IndexStructure = None
        self.PL_IndexMatch = None
        self.PL_IndexLayer = None
        self.PrL_Number = None
        self.PrL_Name = None
        self.PrL_Comment = None
        self.PrL_Type = None

        self.Nt = None
        self.Ne = None
        self.Nc = None
        self.Nm = None
        self.Nr = None
        self.Ng = None

        self.nr_start = None
        self.nr_stop = None

        self.selectors = {}
        self.material_position = None
        self.sector_position = None
        self.multiplier_cementBitumen = None
        self.new_scrap_cycle = None
        self.endUse_aggregates = None
        self.new_scrap_cycle_cutoff = None  # if set as year of model run, set new scrap cycle after this year to zero

    def check_attributes(self):
        """
        Checks if all attributes of the config object are set.

        Monte Carlo object is allowed to be None if config was parsed without uncertainty values. \
        Note that this does not check if set objects are of the correct type or sensible value.

        Raises:
            AttributeError: If parameter dict or model classifications are empty or an attribute, \
                other than the Monte Carlo object, are not set.
        """
        if len(self.parameter_dict) == 0:
            raise AttributeError("Parameter Dict is empty")
        if len(self.model_classification) == 0:
            raise AttributeError("Model classification is empty")
        for var in self.__slots__:
            if getattr(self, var, None) is None:
                if var in ["_monte_carlo", "sensitivity_settings", "uncertainty_settings", "new_scrap_cycle_cutoff"]:
                    continue
                error_msg = f"MISO Config object contains uninitialised parameter: {var}"
                logger.error(error_msg)
                raise AttributeError(error_msg)

    def _create_index_table(self):
        """
        Creates a dictionary with Aspect, Description, dimensions, classification and index letter.
        """
        logger.info('Define index table dataframe.')
        self.index_table = pd.DataFrame({'Aspect': self.IT_Aspects,  # 'Time' and 'Element' must be present!
                                         'Description': self.IT_Description,
                                         'Dimension': self.IT_Dimension,
                                         'Classification': [self.model_classification[Aspect]
                                                            for Aspect in self.IT_Aspects],
                                         'IndexLetter': self.IT_IndexLetter})
        # Unique one letter (upper or lower case) indices to be used later for calculations.

        # Default indexing of self.index_table, other indices are produced on the fly
        self.index_table.set_index('Aspect', inplace=True)

        # Add indexSize to self.index_table:
        self.index_table['IndexSize'] = pd.Series(
            [len(self.index_table.Classification[i].Items) for i in range(0, len(self.index_table.IndexLetter))],
            index=self.index_table.index)

        # list of the classifications used for each indexletter
        self.index_table_classification_names \
            = [self.index_table.Classification[i].Name for i in range(0, len(self.index_table.IndexLetter))]

        logger.debug(f"Index table classification names: {self.index_table_classification_names}")

    def _create_classifications(self):
        """
        Creates the ODYM classification objects.

        Raises:
            AttributeError: If an item select error was found in datafile.
        """
        logger.info('Define model classifications and select items for model classifications \
                    according to information provided by config file.')
        for m in range(0, len(self.IT_Aspects)):
            self.model_classification[self.IT_Aspects[m]] \
                = deepcopy(self.master_classification[self.IT_Classification[m]])
            EvalString \
                = msf.EvalItemSelectString(
                    self.IT_Selector[m], len(self.model_classification[self.IT_Aspects[m]].Items))
            if EvalString.find(':') > -1:  # range of items is taken
                RangeStart = int(EvalString[0:EvalString.find(':')])
                RangeStop = int(EvalString[EvalString.find(':') + 1::])
                self.model_classification[self.IT_Aspects[m]].Items \
                    = self.model_classification[self.IT_Aspects[m]].Items[RangeStart:RangeStop]
            elif EvalString.find('[') > -1:  # selected items are taken
                self.model_classification[self.IT_Aspects[m]].Items \
                    = [self.model_classification[self.IT_Aspects[m]].Items[i] for i in eval(EvalString)]
            elif EvalString == 'all':
                pass
                # not implemented
            else:
                logger.error('Item select error for aspect ' + self.IT_Aspects[m] + ' were found in datafile.')
                raise AttributeError(self)

    def _define_dimension_sizes(self):
        """
        Defines model dimensions.

        Set the time, element, region, material and classification limits of the model as integers.
        """
        # Define dimension sizes
        logger.info("Defining dimension sizes")
        self.Nt = len(self.index_table.Classification[self.index_table.index.get_loc('Time')].Items)
        self.Ne = len(self.index_table.Classification[self.index_table.index.get_loc('Element')].Items)
        self.Nc = len(self.index_table.Classification[self.index_table.index.get_loc('Age-cohort')].Items)
        self.Nm = len(self.index_table.Classification[self.index_table.index.get_loc('Material')].Items)
        self.Nr = len(self.index_table.Classification[self.index_table.index.get_loc('Region')].Items)
        self.Ng = len(self.index_table.Classification[self.index_table.index.get_loc('End-Use Sectors')].Items)
        logger.debug(f"Set Values Nt: {self.Nt}, Ne: {self.Ne}, Nc: {self.Nc}, Nm: {self.Nm}, Nr: {self.Nr}, \
                          Ng: {self.Ng}")

    def initialise_from_excel(self, config_filename, classification_filename,
                              parse_uncertainty=True,
                              uncertainty_settings_filename="MISO2_uncertainty_distribution_settings.xlsx"):
        """
        Parses MISO/ODYM data from XLSX files into the config to configure it.

        The initialisation makes heavy use of the ODYM file format and parsing routines. \
            See the relevant documentation there for more details.

        Arguments:
            config_filename(str): Name of the config.xls.
            classification_filename(str): Name of the master classification xls.
            parse_uncertainty(bool/str): Try to parse uncertainty data from the parameter files. Defaults to true.
            uncertainty_settings_filename(str): Name of the parameter limits xls. Defaults to \
                "MISO2_uncertainty_settings.xls".

        Raises:
            AttributeError: If any config attribute has not been set to some value.
            WasteRateError: If the combined recov and unrecov waste rates of the parsed data \
                exceed 1. This is an error in the input data.

        """
        logger.info("Initialising config object from excel")
        logger.debug(f"Config filename: {config_filename}, Classification filename: {classification_filename}")

        self._load_projects_spec(config_filename)
        self._load_classifications(classification_filename)
        self._set_model_duration()
        self._create_index_table()
        self._define_dimension_sizes()

        self._read_uncertainty_settings_from_excel(os.path.join(self.data_path, uncertainty_settings_filename))

        self._read_parameterdict_from_excel(parse_uncertainty)

        self._read_additional_properties()
        self._set_additional_properties()

        self.nr_start = 0
        self.nr_stop = self.Nr

        if parse_uncertainty == "cv":
            logger.info("Parsing uncertainty data and creating Monte Carlo objects")
            self._monte_carlo = MISO2SensitivityByParameter(self.sensitivity_settings)
            self._monte_carlo.transform_uncert_array(self.parameter_dict, self.uncertainty_settings)
        elif parse_uncertainty:
            self._monte_carlo = MISO2MonteCarlo()
            self._monte_carlo.set_parameter_names(self.PL_Names)
            transform_stats_array_to_dicts(self.parameter_dict, self.uncertainty_settings)

        # Error checking ###

        self.check_attributes()
        preprocessing.checks.MISO2_checks.check_waste_ratios(self.parameter_dict)
        preprocessing.checks.MISO2_checks.check_buffer_year(self.parameter_dict)

    def get_bias(self, mode):
        """
        Returns bias configuration of underlying monte carlo object, if any. Else returns "no uncertainty".

        Args:
            mode(str): One of "next" or "last". Indicates whether next sample bias is returned, or previous sample bias.
        Returns:
            bias(str):
        """
        if self._monte_carlo is not None:
            return self._monte_carlo.get_bias(mode)

        return "no uncertainty"

    def set_monte_carlo_mode(self, mode, **kwargs):
        """
        Set monte carlo method to one of "normal", "systematic_bias", "systematic_bias_v2" or "sensitivity_by_parametergroup"

        Args:
            mode(str): String of mode

        Raises:
            AttributeError: If mode is not known.
        """
        if mode == "normal":
            logger.info("Setting MC mode normal")
            self._monte_carlo = MISO2MonteCarlo()
            self._monte_carlo.set_parameter_names(self.PL_Names)
        elif mode == "systematic_bias":
            logger.info("Setting MC mode systematic bias")
            self._monte_carlo = MISO2MonteCarloSystematic()
            self._monte_carlo.set_parameter_names(self.PL_Names)
        elif mode == "systematic_bias_v2":
            self._monte_carlo = MISO2MonteCarloSystematic_v2()
            self._monte_carlo.set_parameter_names(self.PL_Names)
        elif mode == "sensitivity_by_parametergroup":
            self._monte_carlo = MISO2SensitivityByParameter(self.sensitivity_settings, kwargs["cv_factor"])
        else:
            raise AttributeError(f"Unknown mc mode {mode}")

    def _read_additional_properties(self, data_path=None,
                                    classification_properties_filename="MISO2_classification_properties.xlsx"):
        """
        Parse additional properties from master classification.

        Creates a dictionary of additional properties from the master classification file.
        This functionality seems to be missing from the ODYM parsing.

        Args:
            data_path(None/str): Path to data. If None (default), gathered from miso config class variable.
            classification_properties_filename(str): Path to the classification file.
        """

        if not data_path:
            data_path = self.data_path

        classification_properties_path = os.path.join(data_path, classification_properties_filename)

        if not os.path.exists(classification_properties_path):
            logger.error(f"No such file {classification_properties_path}")
            raise FileNotFoundError

        logger.info(f"Parsing additional properties from {classification_properties_path}")

        properties = pd.read_excel(classification_properties_path, sheet_name=None)

        non_classification_properites = [InputConstants.MULTIPLIER_CEMENT_BITUMEN]

        for k, df in properties.items():
            if k not in non_classification_properites:
                df = df.set_index(InputConstants.CLASSIFICATION_INDEX_COLUMN)
                matching_classification = self.master_classification[k]
                matching_classification.AdditionalProps = df.to_dict(orient='index')

                no_items = len(matching_classification.Items)
                no_props = len(matching_classification.AdditionalProps)

                non_matching_items = set(
                    matching_classification.AdditionalProps.keys()).symmetric_difference(
                    set(matching_classification.Items))

                if no_items != no_props:
                    logger.warning(f"For classification object {k}, parsed a different number of \
                                   additional properties ({no_props}) than items in the classification \
                                       object ({no_items}). \n \
                                   This is likely caused by a change in the MISO2_master_classification \
                                       that is not reflected in the MISO2_additional_properties file")
                if non_matching_items:
                    logger.warning(f"For classification object {k}, a number of items is \
                                   missing from either master classification or additional \
                                       properties parsing: {non_matching_items}")

        multipliers = properties[InputConstants.MULTIPLIER_CEMENT_BITUMEN]
        multipliers = multipliers.set_index(InputConstants.CLASSIFICATION_INDEX_COLUMN)
        multipliers = multipliers.to_dict(orient='index')
        for material_name, multiplier_dict in multipliers.items():
            self.master_classification[InputConstants.CLASSIFICATION_MATERIAL_SHEET].AdditionalProps[material_name][
                InputConstants.MULTIPLIER_CEMENT_BITUMEN] = multiplier_dict

    def get_parameter_dict(self, nr_start=None, nr_stop=None):
        """
        Returns a copy of the ODYM-Parameter dictionary.

        If the config contains multiple regions, the method will return a subset copy by the configs nr_start
        and nr_stop indices. When Monte Carlo randomization is enabled, the underlying MISO2MonteCarlo object
        will create a queue with randomized versions of the Values. If there is only one region in the config
        and Monte Carlo randomization is disabled, it will return a reference.

        Please note that this method does not support returning randomized parameter dicts of more than one region.

        Arguments:
            nr_start(int): Region start index. Defaults to configs nr_start.
            nr_stop(int): Region stop index. Defaults to configs nr_stop.

        Raises:
            WasteRatioError: if the recov and unrecov waste ratio of a process exceed 1.
            InputLimitError: if the values violate the specified hard limits.
        """

        logger.info("Getting parameter dict")

        if nr_start is None:
            nr_start = self.nr_start
        if nr_stop is None:
            nr_stop = self.nr_stop

        retries = 3
        # number of retries for randomisation, if randomised values violate model constraints

        if nr_start == 0 and nr_stop == 1:
            parameter_dict = self.parameter_dict
            # no subset needed, return original (which might get randomised)

        else:
            parameter_dict = misof.copy_and_subset_parameter_dict(
                self.parameter_dict, nr_start, nr_stop)

        randomization_active = self.get_randomization_state()

        if randomization_active:
            if (nr_stop - nr_start) != 1:
                error_msg = "You tried to generate a randomized version of a parameter dict with more than one" \
                            "region. This behaviour is not supported due to likely memory limitations and" \
                            "indexing issues. If you want to use Monte Carlo randomization, split the config into" \
                            "one-region configs first using the $split_config method."
                logger.error(error_msg)
                raise ValueError(error_msg)
            logger.info("Returning monte carlo version of parameter dict")
            try:
                logger.info("Randomization active")
                return_parameter_dict = self._monte_carlo.randomize_model_input(
                    parameter_dict=parameter_dict,
                    uncertainty_settings=self.uncertainty_settings,
                    sample_size=self.mc_batch_size)

                preprocessing.checks.MISO2_checks.check_waste_ratios(return_parameter_dict)
                preprocessing.checks.MISO2_checks.check_limits(return_parameter_dict, self.uncertainty_settings)
                # need to check here since it may trigger retry
            except (WasteRateError, InputLimitError) as error:
                logger.warning(f"During MC randomization the following error was encountered: {repr(error)}. Retrying")
                retries = retries - 1
                if retries:
                    return_parameter_dict = self._monte_carlo.randomize_model_input(
                        parameter_dict, self.uncertainty_settings)

                    preprocessing.checks.MISO2_checks.check_waste_ratios(return_parameter_dict)
                    preprocessing.checks.MISO2_checks.check_limits(return_parameter_dict, self.uncertainty_settings)

                else:
                    logger.error(f"During MC randomization the following error was encountered: {repr(error)}. \
                                      It could not be fixed after several retries ")
                    raise
        else:
            logger.info("Returning non-randomized version of parameter dict")
            return_parameter_dict = parameter_dict
            preprocessing.checks.MISO2_checks.check_waste_ratios(return_parameter_dict)
            preprocessing.checks.MISO2_checks.check_limits(return_parameter_dict, self.uncertainty_settings)

        return return_parameter_dict

    def _set_additional_properties(self):
        """
        Sets additional properties as read from the classification properties file.

        The method will set in-place multipliers for cement bitumen, endUse aggregate \
            multipliers and selectors for aggregates, recycling and downcycling.
        """

        logger.info("Setting additional properties")

        no_materials = self.index_table.loc["Material"]["IndexSize"]

        self.material_position = dict(zip(
            self.index_table.Classification[self.index_table.index.get_loc(
                'Material')].Items, list(range(0, no_materials))))

        no_sectors = self.index_table.loc["End-Use Sectors"]["IndexSize"]
        self.sector_position = dict(zip(
            self.index_table.Classification[self.index_table.index.get_loc(
                'End-Use Sectors')].Items, list(range(0, no_sectors))))

        self.multiplier_cementBitumen = np.ones((len(self.material_position),
                                                 len(self.sector_position)))
        # first dimension is for yet unused element property

        self.new_scrap_cycle = np.ones((len(self.material_position)))

        endUse_aggregates = {}
        selector_Aggr = []
        selector_NoAggr = []
        selectors_downc_mats = []
        multiplier_materials = []

        for material_name, additional_props in self.master_classification[
                InputConstants.CLASSIFICATION_MATERIAL_SHEET].AdditionalProps.items():
            material_pos = self.material_position[material_name]

            self.new_scrap_cycle[material_pos] = float(additional_props["New_scrap_cycle"])

            for sector_name, multiplier in additional_props[InputConstants.MULTIPLIER_CEMENT_BITUMEN].items():
                sector_pos = self.sector_position[sector_name]
                self.multiplier_cementBitumen[material_pos, sector_pos] = float(multiplier)

            if additional_props["Aggregate"]:
                selector_Aggr.append(material_pos)
            else:
                selector_NoAggr.append(material_pos)

            if additional_props["Downc"]:
                selectors_downc_mats.append(material_pos)

            if additional_props["Multiplier_material"]:
                multiplier_materials.append(material_pos)

        # ! NoAggr need to be positioned before Aggr for model core to work
        # (because aggregates are estimated based on use of bitumen/asphalt, bricks, cement/concrete)
        separator_OthAggr = [selector_NoAggr, selector_Aggr]

        for end_use_name, additional_props in self.master_classification["End_use_sectors"].AdditionalProps.items():
            aggr_multiplier = additional_props["aggregate_multiplier"]
            if aggr_multiplier == "NoAggregateMultiplier":
                endUse_aggregates[end_use_name] = None
            else:
                endUse_aggregates[end_use_name] = aggr_multiplier

        self.endUse_aggregates = endUse_aggregates
        self.selectors["selector_Aggr"] = selector_Aggr
        self.selectors["selector_NoAggr"] = selector_NoAggr
        self.selectors["separator_OthAggr"] = separator_OthAggr

        self.selectors["selector_downc_mats"] = selectors_downc_mats
        self.selectors["multiplier_materials"] = multiplier_materials

        logger.debug(f"Set selectors aggregate {selector_Aggr}")
        logger.debug(f"Set selectors no aggregate {selector_NoAggr}")
        logger.debug(f"Set selectors foundation multiplier materials {multiplier_materials}")
        logger.debug(f"Set multiplier cement bitumen {self.multiplier_cementBitumen} ")
        logger.debug(f"Set end use aggregates {endUse_aggregates}")
        logger.debug(f"Set separator other aggregate {separator_OthAggr}")
        logger.debug(f"New scrap cycle {self.new_scrap_cycle}")

    def save_to_pickle(self, filename=None, folder=None):
        """
        Saves the entire MISO Config object to a pickle.

        Args:
            filename(str): Name of the file. Defaults to <UUID>_config.pickle
            folder(str): Folder path. Defaults to config_path variable of MISO config.
        """

        logger.info("Saving MISO Config object to pickle file: ")

        if folder is None:
            folder = self.config_path
            logger.info("No folder path specified, saving to default folder: " + folder)

        if filename is None:
            filename = str(self.unique_id) + "_config.pickle"
            logger.info("No file name specified, saving to default filename (UUID): " + filename)
        with open(os.path.join(folder, filename), 'wb') as handle:
            pickle.dump(self, handle, protocol=pickle.HIGHEST_PROTOCOL)

    def set_global_monte_carlo_state(self, new_state, parameter_list=None):
        """
        Set Monte Carlo randomization of all parameters to the new state.

        This will reset the MISO2MonteCarlo objects queue of randomized dicts (if there was one).

        Args:
            new_state(bool): Set Monte Carlo to on/off.
            parameter_list(list): Optional: Specify parameters to be set as a list of strings.

        Raises:
            AttributeError: when no Monte Carlo object is set in the config.
        """

        logger.info(f"Setting new Monte Carlo state: {new_state}")

        if self._monte_carlo is None:
            error_msg = "No Monte Carlo object found in config. Most likely, the config was not parsed with" \
                        "parse_uncertainty = True."
            logger.error(error_msg)
            raise AttributeError(error_msg)

        if parameter_list is None:
            logger.info("No parameter list provided, setting for all values that"
                        "have uncertainty according to distribution settings")
            parameter_list = []
            for parameter_name, values_dict in self.uncertainty_settings.items():
                if isinstance(values_dict["Uncertainty_values"], str):
                    parameter_list.append(parameter_name)

        self._monte_carlo.update_parameter_randomization(new_state, parameter_list)

    def get_randomization_state(self):
        """
        Get Monte Carlo state of config.

        If a Monte Carlo object exists and any parameter is set to be randomized, returns True. Returns false otherwise.

        Returns:
            mc_state(bool): Monte Carlo state of config.

        """
        logger.info("Checking Configs randomization state")

        if self._monte_carlo is None:
            return False

        return self._monte_carlo.get_parameter_randomization()

    def _load_projects_spec(self, config_filename):
        """
        Load project specification and parse it with :func:`ODYMs parse model control \
            <ODYM_Functions.ParseModelControl>`. Data path will be taken from config data.

        Args:
            config_filename(str): Name of the config.
        """

        logger.info("Loading Project Specs")
        model_configfile = openpyxl.load_workbook(os.path.join(self.data_path, config_filename), data_only=True)
        self.script_config = {'Model Setting': model_configfile['Config'].cell(4, 4).value}
        model_configsheet = model_configfile[('Setting_' + self.script_config['Model Setting'])]
        name_scenario = model_configsheet.cell(4, 4).value
        logger.info("Scenario name:" + name_scenario)
        self.script_config = msf.ParseModelControl(model_configsheet, self.script_config)

        logger.info('Reading and parsing config table, including the model index table, from model config sheet.')
        IT_Aspects, IT_Description, IT_Dimension, IT_Classification, IT_Selector, IT_IndexLetter, \
            PL_Names, PL_Description, PL_Version, PL_IndexStructure, PL_IndexMatch, PL_IndexLayer, \
            PrL_Number, PrL_Name, PrL_Comment, PrL_Type, self.script_config \
            = msf.ParseConfigFile(model_configsheet, self.script_config, logger)

        self.IT_Aspects = IT_Aspects
        self.IT_Description = IT_Description
        self.IT_Dimension = IT_Dimension
        self.IT_Classification = IT_Classification
        self.IT_Selector = IT_Selector
        self.IT_IndexLetter = IT_IndexLetter
        self.PL_Names = PL_Names
        self.PL_Description = PL_Description
        self.PL_Version = PL_Version
        self.PL_IndexStructure = PL_IndexStructure
        self.PL_IndexMatch = PL_IndexMatch
        self.PL_IndexLayer = PL_IndexLayer
        self.PrL_Number = PrL_Number
        self.PrL_Name = PrL_Name
        self.PrL_Comment = PrL_Comment
        self.PrL_Type = PrL_Type

        logger.debug("Set specs: ")
        logger.debug(f"IT Aspects: {IT_Aspects} ")
        logger.debug(f"IT Description: {IT_Description}")
        logger.debug(f"IT Dimension: {IT_Dimension} ")
        logger.debug(f"IT Classification: {IT_Classification} ")
        logger.debug(f"IT Selector: {IT_Selector}")
        logger.debug(f"IT Index Letter: {IT_IndexLetter}")
        logger.debug(f"PL Names: {PL_Names}")
        logger.debug(f"PL Description: {PL_Description}")
        logger.debug(f"PL Version: {PL_Version}")
        logger.debug(f"PL Index Structure: {PL_IndexStructure}")
        logger.debug(f"PL IndexMatch: {PL_IndexMatch}")
        logger.debug(f"PL Index Lager: {PL_IndexLayer} ")
        logger.debug(f"Prl Number: {PrL_Number}")
        logger.debug(f"Prl Name: {PrL_Name} ")
        logger.debug(f"PrL Comment: {PrL_Comment}")
        logger.debug(f"PrL Type: {PrL_Type}")
        logger.debug(f"Script config: {self.script_config}")

    def _load_classifications(self, classification_filename):
        """
        Load the classification file from excel and parse it with the ODYM classification parser.

        Args:
            classification_filename(str): Filename of classification within the config's data_path.
        """

        classification_file = openpyxl.load_workbook(
            os.path.join(self.data_path, classification_filename), data_only=True)
        classification_sheet = classification_file['MAIN_Table']
        self.master_classification = msf.ParseClassificationFile_Main(classification_sheet, logger)
        self._create_classifications()

    def _read_parameterdict_from_excel(self, parse_uncertainty):
        """
        Read a parameter dict from XLSX.

        This method utilities ODYM parsing functions to create a parameter dictionary \
            from Excel files in the ODYM format. May raise exceptions from within the ODYM code.

        Args:
            parse_uncertainty(bool): If uncertainty data should be parsed or not.
        """

        logger.info('Reading parameters from excel')

        mo_start = 0
        for mo in range(mo_start, len(self.PL_Names)):
            parameter_dict_path = os.path.join(self.data_path, self.PL_Names[mo] + '_' + self.PL_Version[mo])
            logger.info(f'Reading parameter {self.PL_Names[mo]}')
            # Do not change order of parameters handed over to function!
            if parse_uncertainty:
                sheet_name = "cv" if parse_uncertainty == "cv" else None

                logger.info("Parsing with uncertainty = true")
                MetaData, Values, Uncertainty = msf.ReadParameterXLSX(
                    parameter_dict_path, self.PL_Names[mo], self.PL_IndexStructure[mo], self.PL_IndexMatch[mo],
                    self.PL_IndexLayer[mo], self.master_classification, self.index_table,
                    self.index_table_classification_names, self.script_config, logger, True, sheet_name)

                self.parameter_dict[self.PL_Names[mo]] = msc.Parameter(
                    Name=MetaData['Dataset_Name'], ID=MetaData['Dataset_ID'],
                    UUID=MetaData['Dataset_UUID'], P_Res=None, MetaData=MetaData,
                    Indices=self.PL_IndexStructure[mo], Values=Values, Uncert=Uncertainty,
                    Unit=MetaData['Dataset_Unit'])
            else:
                MetaData, Values = msf.ReadParameterXLSX(
                    parameter_dict_path, self.PL_Names[mo], self.PL_IndexStructure[mo], self.PL_IndexMatch[mo],
                    self.PL_IndexLayer[mo], self.master_classification, self.index_table,
                    self.index_table_classification_names, self.script_config, logger, parse_uncertainty)

                self.parameter_dict[self.PL_Names[mo]] = msc.Parameter(
                    Name=MetaData['Dataset_Name'], ID=MetaData['Dataset_ID'],
                    UUID=MetaData['Dataset_UUID'], P_Res=None, MetaData=MetaData,
                    Indices=self.PL_IndexStructure[mo], Values=Values,
                    Unit=MetaData['Dataset_Unit'])
                logger.debug("Current parameter file Dataset UUID: " + MetaData['Dataset_UUID'])

    def _read_uncertainty_settings_from_excel(self, uncertainty_settings_filename=None):
        """
        Read the parameter limits from Excel.

        This additional parsing step is necessary since ODYM does not allow for parsing additional information of parameter uncertainty.
        We decided not to modify the ODYM functionality to retain compatibility.

        Args:
            uncertainty_settings_filename(str): Filename of parameter limits.\
                Defaults to Filename in InputConstants if None.

        Returns:
            uncertainty_settings_dict(dict): Dictionary of parameter names and Min,Max and Type information.
        """
        if uncertainty_settings_filename is None:
            uncertainty_settings_filename = InputConstants.UNCERTAINTY_DISTRIBUTION_FILENAME

        logger.info(f"Reading uncertainty settings from: {uncertainty_settings_filename}")

        all_uncertainty_settings = pd.read_excel(uncertainty_settings_filename, sheet_name=None)

        uncertainty_settings = all_uncertainty_settings["main"]

        self.uncertainty_settings = uncertainty_settings.set_index("Parameter_Name").T.to_dict()

        if "sensitivity" in all_uncertainty_settings.keys():
            self.sensitivity_settings = all_uncertainty_settings["sensitivity"] # .to_dict()
        # return uncertainty_settings_dict

    def _set_model_duration(self):
        """
        Set default values for model duration.

        These will the lowest and highest value of the model_classification["Time"] entry, with duration as their
        difference + 1.
        """

        logger.info('Defining model index table and parameter dictionary')
        self.model_time_start = int(min(self.model_classification['Time'].Items))
        self.model_time_end = int(max(self.model_classification['Time'].Items))
        self.model_duration = self.model_time_end - self.model_time_start + 1
        logger.debug(f"Model duration: {self.model_duration}")

    def set_random_state(self, random_state):
        """
        Set a seed for the Monte Carlo randomization to get reproducible results.

        If this is set to any integer value, the Monte Carlo randomization will always return the same results.
        !!!: This is currently only working for one batch of randomized data inputs.

        Args:
            random_state(int): Seed that is passed to the Scipy distributions.
        """

        logger.info(f"Setting Monte Carlo seed: {random_state}")
        self._monte_carlo.set_random_state(random_state)

    def reload_classification_properties(self,
                                         classification_properties_filename="MISO2_classification_properties.xlsx"):
        """
        Reload the classification properties from file.

        Use this method to reload the classification properties if they have changed
        after initialisation of the config object.

        Warning: Additional props are reset to None before parsing the file again, so any
        exception in the setting of classifications will leave the miso config in an illegal state.

        Args:
            classification_properties_filename(str) = Name of the properties filename in the
                configs data path. Defaults to "MISO2_classification_properties.xlsx"

        """

        logger.info(f"Reloading classification properties from {classification_properties_filename}")

        for classification in self.master_classification.values():
            classification.AdditionalProps = None

        self._read_additional_properties(data_path=None,
                                         classification_properties_filename=classification_properties_filename)
        self._set_additional_properties()

    def get_new_scrap_cycle(self, time):
        """
        Returns a copy of the new scrap cycle array.
        If there is a cutoff value set in the config, will return all zeros (scenario option).

        Args:
            time(int): Time index in years that is compared to cutoff value.

        Returns:
            new_scrap_cycle(np.array): Array of new scrap cycle
        """
        if self.new_scrap_cycle_cutoff is not None and time > self.new_scrap_cycle_cutoff:
            logger.info(f"Cutoff year {self.new_scrap_cycle_cutoff} reached, returning zero for new scrap")
            return np.zeros_like(self.new_scrap_cycle)
        else:
            return self.new_scrap_cycle.copy()

    def save_metadata(self, output_path, additional_data_dir):
        """
        Saves metadata to xlsx.

        Note that not all metadata is constructed automatically, check the file to make sure it is correct.

        Args:
            output_path: Path to save output file to.
            additional_data_dir: Directory where Metadata template is located.
        """

        metadata = {"MISO2_software_version": version('MISO2'),
                    "MISO2_database_version": self.script_config["Version of master classification"],
                    "Date_created": datetime.now().strftime("%Y_%m_%d_%H:%M:%S"),
                    "MISO2_config_id": str(self.unique_id), "Units": "kilotons", "Comment": ""}

        dimensions_metadata = {"Countries": sorted(self.master_classification["Countries"].Items),
                               "Materials": sorted(self.master_classification["SB_Material"].Items),
                               "End-uses": sorted(self.master_classification["End_use_sectors"].Items)}
        all_years = self.master_classification["Time"].Items
        dimensions_metadata["Time"] = (min(all_years), max(all_years))

        metadata_file_path_input = os.path.join(additional_data_dir, InputConstants.METADATA_FILENAME)
        metadata_file_path_output = os.path.join(output_path, "MISO2_metadata.xlsx")
        shutil.copyfile(metadata_file_path_input, metadata_file_path_output)
        with pd.ExcelWriter(metadata_file_path_output, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            pd.DataFrame.from_dict([metadata]).T.to_excel(writer, sheet_name="Metadata", index=True, header=False)
            pd.DataFrame.from_dict(
                dimensions_metadata, orient="index").T.to_excel(
                writer, sheet_name="Dimensions", index=False, header=True)


def split_config(miso_config, index):
    """
    Returns a subset of the config by region index.

    As far as possible, it will give views of the data instead of copies. Split objects receive new ids.
    This subset config is mainly intended to be passed to another process, by copy.

    Args:
        miso_config(MISO2Config): Config to be subset.
        index(int): Regional index number to subset. Must be within the index range of the MISO2Config.

    Returns:
        split_config(MISO2Config): Config subset to one region.
    """

    logger.info(f"Splitting MISO Config index: {index}")

    new_split_config = MISO2Config(config_path=miso_config.config_path,
                                   data_path=miso_config.data_path,
                                   timestamp=miso_config.timestamp)

    new_split_config.model_classification = deepcopy(miso_config.model_classification)
    new_split_config.model_classification["Region"].Items = miso_config.model_classification["Region"].Items[index]
    new_split_config.uncertainty_settings = miso_config.uncertainty_settings
    new_split_config.unique_id = uuid.uuid4()

    new_split_config.parameter_dict = {}

    for k, v in miso_config.parameter_dict.items():

        values_view = v.Values[index, ...]
        values_view = values_view[np.newaxis, ...]
        split_uncert = None
        if v.Uncert is not None:
            if miso_config.uncertainty_settings[k]["Type"] == 'Allocation':
                split_uncert = v.Uncert[index, ...]
                split_uncert= split_uncert[np.newaxis, ...]
            else:
                uncert_copy = np.array(v.Uncert).reshape(v.Values.shape)[index, ...]
                uncert_copy = uncert_copy[np.newaxis, ...]
                split_uncert = uncert_copy.flatten().tolist()
            # to list implicit copy call, not sure how to prevent that

        new_param = msc.Parameter(Name=v.Name, ID=v.ID, UUID=v.UUID,
                                  P_Res=v.P_Res, MetaData=v.MetaData, Indices=v.Indices,
                                  Values=values_view,
                                  Uncert=split_uncert)
        new_split_config.parameter_dict[k] = new_param

    new_split_config.mc_batch_size = miso_config.mc_batch_size
    new_split_config._monte_carlo = deepcopy(miso_config._monte_carlo)
    new_split_config.script_config = miso_config.script_config

    # Master Classification
    new_split_config.master_classification = deepcopy(miso_config.master_classification)
    new_split_config.master_classification["Countries"].Items \
        = [miso_config.master_classification["Countries"].Items[index]]

    # Index Table
    new_split_config.index_table = miso_config.index_table.copy(deep=True)
    new_split_config.index_table.loc["Region", "IndexSize"] = 1
    # for this, deepcopy works

    classification_list = deepcopy(list(miso_config.index_table.Classification))

    new_split_config.index_table.Classification = classification_list

    # pandas deepcopy does not copy nested objects
    # workaround here necessary due to ODYM data structure

    new_split_config.index_table.Classification.loc["Region"].Items \
        = [new_split_config.index_table.Classification.loc["Region"].Items[index]]
    new_split_config.index_table_classification_names = miso_config.index_table_classification_names

    # Other
    new_split_config.model_duration = miso_config.model_duration
    new_split_config.model_time_start = miso_config.model_time_start
    new_split_config.model_time_end = miso_config.model_time_end
    new_split_config.IT_Aspects = miso_config.IT_Aspects

    new_split_config.IT_Description = miso_config.IT_Description
    new_split_config.IT_Dimension = miso_config.IT_Dimension
    new_split_config.IT_Classification = miso_config.IT_Classification
    new_split_config.IT_Selector = miso_config.IT_Selector
    new_split_config.IT_IndexLetter = miso_config.IT_IndexLetter

    new_split_config.PL_Names = miso_config.PL_Names
    new_split_config.PL_Description = miso_config.PL_Description
    new_split_config.PL_Version = miso_config.PL_Version
    new_split_config.PL_IndexStructure = miso_config.PL_IndexStructure
    new_split_config.PL_IndexMatch = miso_config.PL_IndexMatch
    new_split_config.PL_IndexLayer = miso_config.PL_IndexLayer
    new_split_config.PrL_Number = miso_config.PrL_Number
    new_split_config.PrL_Name = miso_config.PrL_Name
    new_split_config.PrL_Comment = miso_config.PrL_Comment
    new_split_config.PrL_Type = miso_config.PrL_Type

    new_split_config.Nt = miso_config.Nt
    new_split_config.Ne = miso_config.Ne
    new_split_config.Nc = miso_config.Nc
    new_split_config.Nm = miso_config.Nm
    new_split_config.Nr = 1
    new_split_config.Ng = miso_config.Ng

    new_split_config.nr_start = 0
    new_split_config.nr_stop = 1

    new_split_config.selectors = miso_config.selectors.copy()

    new_split_config.material_position = miso_config.material_position.copy()
    new_split_config.sector_position = miso_config.sector_position.copy()
    new_split_config.endUse_aggregates = miso_config.endUse_aggregates.copy()
    new_split_config.multiplier_cementBitumen = miso_config.multiplier_cementBitumen.copy()
    new_split_config.new_scrap_cycle = miso_config.new_scrap_cycle.copy()

    new_split_config.sensitivity_settings = miso_config.sensitivity_settings.copy()

    new_split_config.check_attributes()

    return new_split_config


def load_configs_from_folder(config_path, key="id"):
    """
    Tries to load all .pickle files in a directory as MISO2Configs and returns them as a list.

    Args:
        config_path(str): A filepath
        key(str): Key of return dict. One of "id"(default) or filename.

    Returns:
         configs(dict): A dict of MISO2Config objects.
    """

    restored_configs = {}
    config_filenames = glob.glob(os.path.join(config_path, "*.pickle"))

    for filename in config_filenames:
        try:
            split_config_path = os.path.join(config_path, filename)
            with open(split_config_path, 'rb') as f:
                miso_config = pickle.load(f)
            if key == "id":
                restored_configs[miso_config.unique_id] = miso_config
            elif key == "filename":
                restored_configs[config_path] = miso_config
            else:
                raise ValueError

        except (IOError, ValueError) as error:
            logger.error("Something happened while trying to load a config")
            logger.exception(error)
    return restored_configs


def save_split_configs_to_folder(path, config_name, config):
    """
    Splits a Config object into multiple objects, one per region, and saves them to a folder.

    Individual files receive a running index.

    Args:
         path(str): Folder path to save to
         config_name(str): Filename of the config.
         config(MISO2Config): Config to split
    """
    end = config.nr_stop
    for i in range (0, end):
        new_filename = "split_" + str(i) + "_" + config_name + ".pickle"
        logger.info(f"split config {new_filename}")
        new_config = split_config(config, i)
        new_config.save_to_pickle(filename=new_filename, folder=path)
